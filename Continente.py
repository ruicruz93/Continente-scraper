# -*- coding: utf-8 -*-

import re, bs4, requests, subprocess, openpyxl, sys, traceback, csv, winsound, random
from openpyxl.styles import colors, fills, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def get_page(url):
    my_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:83.0) Gecko/20100101 Firefox/83.0'}
    page = requests.get(url, headers=my_header)
    page.raise_for_status()
    page_parsed = bs4.BeautifulSoup(page.text, 'lxml')
    page.close()
    return page_parsed


def main():
    
    def link_builder(expression):
        if expression not in groceries_dic.keys():
            return f'pesquisa/?q={"+".join(expression.split())}&'
        else:
            return f'{groceries_dic[expression]}?'     
    
    directory = re.compile(r'(.*)\\[^\\]*$').search(sys.argv[0]).group(1)
    #Get variables paths
    with open(fr'{directory}\Variables.txt', 'r', encoding='utf-8', newline='') as txt:
        reader = csv.DictReader(txt, delimiter=',')
        fieldnames = reader.fieldnames
        variables = tuple(reader)
    for row in variables:
        if row['Path'] == '':
            print('Missing variable paths. Please fill out Variables.txt first.')
            return
    
    #Make groceries dictionary from Categories.txt
    with open(fr'{directory}\Categories.txt', 'r', encoding='utf-8', newline='') as txt:
        reader = csv.DictReader(txt, delimiter=',')
        groceries_dic = {tuple(row.values())[0]:tuple(row.values())[1] for row in reader}
    
    #Unit conversion factor dictionary
    converter_dic = {'un': 1, 'gr': 1000, 'kg': 1, 'ml': 1000, 'cl': 100, 'lt': 1, 'rolos': 1}
    
    print('The following product categories will be scraped for the latest price discounts in Continente:\n\n')
    print('---> ', end='')
    print(*tuple(groceries_dic.keys()), sep='\n---> ')
    
    shopping_list = []
    while True:
        hit = input('\nTo look for another deal, please type in the product\'s name or enter 0 to begin scraping prices\n')
        if hit != '0':
            shopping_list.append(hit)
        else:
            print('Looking for the best deals. You will hear a beep once the process is over, so make sure to leave your speakers unmuted. Feel free to keep using your computer in the meantime...\n\n')
            break
    shopping_list += groceries_dic.keys()
    
    wb = openpyxl.Workbook()
    wb['Sheet'].sheet_properties.tabColor = '00000000' #Setting this now to avoid problems with the random_hex function later on
    
    for p, item in enumerate(shopping_list):
        try:
            item_deals_list = []
            #Each item page can have up to 24 deals, so if it has exactly 24, move to next iteration to check if there's more deals.
            for k in range(100):        
                link = f'https://www.continente.pt/{link_builder(item)}start={k*24}&srule=price-low-to-high&pmin=0.01&prefn1=isPromo&prefv1=true'
                page = get_page(link)
                nr_of_items = len(tuple(i for i in page.select('a.ct-tile--description')))
                if nr_of_items == 0:
                    break                
                names = tuple(i.getText().strip() for i in page.select('a.ct-tile--description'))
                brand = tuple(i.getText().strip() for i in page.select('p.ct-tile--brand'))
                quantity = tuple(re.compile(r'\.?(\d*\.?\d+?) ?(un|gr|kg|cl|ml|lt)', re.I).search(i.getText().strip().replace(',', '.')).groups() if i.getText() != '' else ('1', 'kg') for i in page.select('p.ct-tile--quantity'))
                nu_prices = tuple(re.compile(r'(\d+\.\d+)[^/]*/(\w+)').search(i.getText().replace(',', '.')).groups() for i in page.select('span.sales.ct-tile--price-primary'))
                per_unit_prices = tuple(round(float(nu_prices[i][0]) / float(quantity[i][0]) * converter_dic[quantity[i][1]], 2) if nu_prices[i][1] == 'un' else nu_prices[i][0] for i in range(nr_of_items))
                for l in range(nr_of_items):
                    item_deals_list.append((names[l], brand[l], quantity[l][0] + quantity[l][1], float(nu_prices[l][0]), float(per_unit_prices[l])))
                if nr_of_items in range(1, 24):
                    break
                
            if len(item_deals_list) == 0:
                continue
            
            #Create sheet for item
            wb.create_sheet(index=p, title=item)
            wb[item].freeze_panes = 'A2'
            wb[item].column_dimensions['A'].width = 80
            wb[item].row_dimensions[1].height = 70
            wb[item].sheet_properties.tabColor = '00000000'
            header = ('Name', 'Brand', 'Amount', 'Discount Price', 'Price per Unit|Kg|L')
            
            #Function that will return a random to color to be used as the color for the sheet's tab and header cells
            def random_hex():
                color_chars_tuple = tuple(str(j) for j in range(10)) + tuple(chr(l) for l in range(65, 71))
                color = '00' + ''.join(random.choices(color_chars_tuple, weights=None, k=6)) #random.choices allows duplicates in sample, which is okay here
                if color not in tuple(wb[m].sheet_properties.tabColor.rgb for m in wb.sheetnames):
                    return color
                else:
                    return random_hex()            
            
            nu_sheet_color = random_hex()
            wb[item].sheet_properties.tabColor = nu_sheet_color
            #Fill in header on first row
            for t, j in enumerate(wb[item]['A1':'E1'][0]):
                if t > 0:
                    wb[item].column_dimensions[get_column_letter(t+1)].width = 26.1
                j.value = header[t]
                j.fill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb=nu_sheet_color))
                j.font = Font(name='Times New Roman', size=20)
                j.alignment = Alignment(wrap_text=True)
            item_deals_list.sort(key=lambda x: x[-1]) #Sheets will be sorted by per unit price of deals
            #Fill in values for each deal in a new row
            for a, z in enumerate(item_deals_list):
                wb[item].row_dimensions[a+2].height = 50
                for q in range(5):
                    celula = wb[item].cell(row=a+2, column=q+1)
                    celula.value = z[q]
                    celula.font = Font(name='Times New Roman', size=20)
                    celula.alignment = Alignment(wrap_text=True)
        except:
            print(f"{item} was not scraped due to:\n\n{traceback.print_exc()}\n\n")   
    
    
    del wb['Sheet']
    deals_list = fr'{directory}\Deals List.xlsx'
    wb.save(deals_list)
    wb.close()
    winsound.Beep(440, 2000) #Just to let the user know that the process is over
    subprocess.Popen([variables[0]['Path'], deals_list], shell=False) #Open the newly created deals list
    subprocess.Popen([variables[1]['Path'], 'https://www.continente.pt/login/'], shell=False) #Open the website in case the user wants to order his groceries online

#*************************************************************************************

main()